"""Generate evidence files for the MedTriage AI maturity test scenario."""
from openpyxl import Workbook

# =============================================================================
# 1. Mature AI Policy (txt) - substantively complete but with one gap
# =============================================================================
mature_policy = """MEDTECH HEALTH - AI GOVERNANCE POLICY
Version 2.1 (Approved by Board, March 2025)
Effective Date: 1 April 2025
Next Review: 1 April 2026

1. PURPOSE
This policy establishes the framework for responsible development, deployment,
and operation of AI systems at MedTech Health, with particular focus on systems
affecting patient care and clinical decision support.

2. SCOPE
Applies to all AI systems developed by, deployed by, or procured by MedTech Health,
across all clinical and operational domains. Includes both first-party and
third-party AI systems integrated into MedTech operations.

3. GOVERNANCE STRUCTURE
3.1 The AI Governance Committee, chaired by the Chief Medical Officer, holds
ultimate accountability for AI risk. Committee meets monthly.
Members: CMO (chair), CRO, CISO, Head of Data Science, Head of Clinical Operations,
Head of Legal, Privacy Officer, two clinician representatives.

3.2 The AI Governance Lead is a dedicated 1.0 FTE role reporting to the CRO,
responsible for day-to-day governance, policy maintenance, training coordination,
and external regulatory engagement.

3.3 Each AI system must have a designated Product Owner accountable for
operational compliance. Product Owners report to the AI Governance Committee
quarterly on system performance, incidents, and risk register updates.

4. AI LIFECYCLE GOVERNANCE
4.1 Pre-deployment Impact Assessment: All AI systems require documented impact
assessment using the MedTech AIA template before deployment. Template covers
clinical risks, fundamental rights impacts, fairness considerations, and
data protection.

4.2 Bias and Fairness: All AI systems making decisions affecting patients must
undergo bias testing using the MedTech Bias Testing Methodology v3.0, covering
protected attributes including sex, age, ethnicity, and socioeconomic indicators
where available. Testing performed quarterly in production. Results reviewed
by the AI Governance Committee.

4.3 Human Oversight: All clinical AI systems are designed for human-in-the-loop
operation. AI outputs are recommendations only; final clinical decisions remain
with licensed clinicians. The 'override' rate is monitored as a quality indicator.

4.4 Production Monitoring: All AI systems must have automated monitoring covering:
performance drift, fairness drift, error rates, and prediction distribution.
Alerts route to the AI Governance Lead for triage.

4.5 Transparency: Patients are informed when AI is used in their care via
posted notices and consent forms. Detailed explanations are available on request.

5. THIRD-PARTY AI
5.1 Procurement of third-party AI systems requires AI Governance Committee approval
including review of vendor's bias testing methodology, validation studies, and
ongoing monitoring commitments.

5.2 Annual vendor reassessment is conducted including review of any model updates,
new validation data, and incidents.

6. INCIDENT MANAGEMENT
6.1 Suspected adverse events involving AI systems are reported through the
MedTech Patient Safety Reporting system within 24 hours.

6.2 The AI Governance Lead conducts root cause analysis with the relevant
Product Owner and reports findings to the AI Governance Committee within 14 days.

7. TRAINING
7.1 Mandatory annual AI literacy training for all staff using AI systems.
7.2 Specialized training for AI Governance Committee members and Product Owners.

8. REGULATORY COMPLIANCE
8.1 The AI Governance Committee maintains a regulatory horizon scan including
EU AI Act, FDA guidance, MHRA guidance, and applicable state regulations.

8.2 Note: This policy is currently being updated to address EU AI Act high-risk
system obligations applicable from 2 August 2026. Updated procedures expected by
Q2 2026.

9. POLICY REVIEW
This policy is reviewed annually or upon significant regulatory change.

Approved: Board of Directors, 15 March 2025
"""

with open("medtech_ai_policy.txt", "w") as f:
    f.write(mature_policy)
print("Created medtech_ai_policy.txt")


# =============================================================================
# 2. Bias Testing Methodology (txt) - good methodology
# =============================================================================
bias_methodology = """MEDTECH HEALTH - BIAS TESTING METHODOLOGY
Version 3.0 (Approved by AI Governance Committee, January 2026)

1. PURPOSE
This methodology defines the approach for assessing AI systems at MedTech Health
for fairness and disparate impact across protected attributes.

2. SCOPE
Applies to all AI systems making decisions or recommendations that materially
affect patient care, treatment access, or operational outcomes for individuals.

3. PROTECTED ATTRIBUTES TESTED
3.1 Required attributes (always tested where data available):
- Sex (binary classification per registration data)
- Age (categorized: 18-30, 31-50, 51-70, 71+)
- Ethnicity (where collected with consent per applicable law)
- Socioeconomic indicators (insurance type as proxy where direct data unavailable)

3.2 Additional attributes (tested where ethically appropriate):
- Geographic region (urban/rural)
- Primary language

4. METRICS APPLIED
4.1 Primary metrics:
- Disparate Impact Ratio (target: >= 0.80 across all protected groups)
- Demographic Parity Difference (target: <= 0.05)
- Equal Opportunity Difference (target: <= 0.05)
- Calibration error by group (target: <= 0.02)

4.2 Supplementary metrics applied based on system context:
- False Negative Rate Parity (for safety-critical systems)
- Predictive Value Parity (for treatment recommendation systems)

5. CADENCE
5.1 Pre-deployment baseline test required for all systems.
5.2 Quarterly production retest for clinical AI systems.
5.3 Annual full audit including methodology review.

6. REMEDIATION PROTOCOL
6.1 When metrics fall outside thresholds, AI Governance Committee determines
remediation pathway: model retraining, data rebalancing, threshold adjustment,
or system suspension.

6.2 Remediation must be implemented within 90 days of identification or system
must be suspended pending remediation.

7. DOCUMENTATION
All bias test results, including methodology applied, sample sizes, and any
identified disparities, are documented in the MedTech AI Test Registry.
"""

with open("medtech_bias_methodology.txt", "w") as f:
    f.write(bias_methodology)
print("Created medtech_bias_methodology.txt")


# =============================================================================
# 3. Recent Bias Testing Results (xlsx) - mostly passing, one minor concern
# =============================================================================
wb = Workbook()

ws1 = wb.active
ws1.title = "Q1 2026 Bias Test Results"
ws1.append(["MedTriage AI - Bias Testing Results"])
ws1.append(["Test Date", "2026-04-02"])
ws1.append(["Tested By", "MedTech Data Science Team"])
ws1.append(["Methodology Version", "v3.0"])
ws1.append([])
ws1.append(["Metric", "Sex", "Age", "Ethnicity", "Threshold", "Status"])
ws1.append(["Disparate Impact Ratio", "0.96", "0.91", "0.89", ">= 0.80", "PASS"])
ws1.append(["Demographic Parity Diff", "0.02", "0.03", "0.04", "<= 0.05", "PASS"])
ws1.append(["Equal Opportunity Diff", "0.01", "0.04", "0.06", "<= 0.05", "MARGINAL"])
ws1.append(["Calibration Error", "0.01", "0.02", "0.03", "<= 0.02", "MARGINAL"])

ws2 = wb.create_sheet("By Subgroup")
ws2.append(["Group", "Triage Rate", "Sample Size", "Notes"])
ws2.append(["Male", "82%", "5400", "Reference"])
ws2.append(["Female", "80%", "5200", "Within tolerance"])
ws2.append(["Age 18-30", "78%", "2800", "Within tolerance"])
ws2.append(["Age 31-50", "81%", "3100", "Reference"])
ws2.append(["Age 51-70", "82%", "3400", "Within tolerance"])
ws2.append(["Age 71+", "75%", "1300", "Below reference; monitoring"])
ws2.append(["Ethnicity A", "81%", "4500", "Reference"])
ws2.append(["Ethnicity B", "78%", "3200", "Within tolerance"])
ws2.append(["Ethnicity C", "76%", "2400", "Marginal; remediation in progress"])

ws3 = wb.create_sheet("Remediation Status")
ws3.append(["Issue", "Status", "Owner", "Target Date"])
ws3.append(["Equal Opportunity for Ethnicity C marginal at 0.06", "In remediation", "Data Science Lead", "2026-Q3"])
ws3.append(["Calibration error for Ethnicity C above target", "Investigating", "Data Science Lead", "2026-Q2"])

wb.save("medtech_bias_results_q1_2026.xlsx")
print("Created medtech_bias_results_q1_2026.xlsx")


# =============================================================================
# 4. Impact Assessment (txt) - thorough but one weakness
# =============================================================================
ai_impact_assessment = """MEDTECH HEALTH - AI IMPACT ASSESSMENT
System: MedTriage AI v4.2
Assessor: MedTech AI Governance Lead
Date: 12 January 2026
Approved by AI Governance Committee: 28 January 2026
Next Review: January 2027

1. SYSTEM DESCRIPTION
MedTriage AI is a clinical decision support system that evaluates patient-reported
symptoms in the emergency department waiting area and recommends a triage priority
level (1-5 ESI scale) to assist triage nurses. Final triage decisions remain with
licensed nursing staff. Deployed in 47 hospitals across the EU and UK.

2. INTENDED USE
2.1 Primary use: Reduce nurse workload by pre-screening patient symptoms.
2.2 Decision boundaries: AI provides recommendations only; clinical judgment by
licensed staff is the final determinant of triage priority.
2.3 Out-of-scope: NOT used for definitive diagnosis, treatment selection, or
medication recommendations.

3. RISK ASSESSMENT

3.1 Clinical Risks
- False low-priority assignment leading to delayed care
  Mitigation: Conservative model bias toward over-triage; nurse review mandatory
  Residual risk: LOW

- Model failure leading to recommendation absence
  Mitigation: Manual triage protocol unchanged; AI is augmentation only
  Residual risk: LOW

3.2 Fundamental Rights Risks
- Disparate triage outcomes across demographic groups
  Mitigation: Quarterly bias testing per Methodology v3.0
  Residual risk: MEDIUM (see ongoing remediation for Ethnicity C marginal metric)

- Loss of patient autonomy in triage decisions
  Mitigation: AI recommendations are not binding; informed consent on AI use
  Residual risk: LOW

3.3 Data Protection Risks
- Sensitive health data processed at scale
  Mitigation: All data pseudonymized; lawful basis under Article 9(2)(h) GDPR
  (provision of healthcare). Data Processing Agreement with model vendor.
  Residual risk: LOW

3.4 Operational Risks
- Model drift over time as patient populations change
  Mitigation: Monthly drift monitoring; quarterly retraining cycle
  Residual risk: LOW

4. EU AI ACT CLASSIFICATION
This system is high-risk under EU AI Act Annex III(5)(a) (essential services -
healthcare access). System is being prepared for full EU AI Act compliance by
2 August 2026. Specific actions in progress:
- Updating technical documentation to Article 11 standards
- Establishing post-market monitoring system per Article 72
- Preparing for conformity assessment

5. RECOMMENDATIONS
5.1 Continue current bias testing cadence; complete Ethnicity C remediation by Q3 2026.
5.2 Complete EU AI Act Article 11 documentation by 1 July 2026.
5.3 Annual reassessment scheduled for January 2027.

6. APPROVAL
Recommended for continued operation with mitigations.
Approved by AI Governance Committee, 28 January 2026.
"""

with open("medtech_impact_assessment.txt", "w") as f:
    f.write(ai_impact_assessment)
print("Created medtech_impact_assessment.txt")

print("\nAll mature evidence files created.")